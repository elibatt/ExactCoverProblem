"""
Funzioni Comuni
"""
import gc
import sys
from sys import getsizeof
import numpy as np
import xlsxwriter
from openpyxl import load_workbook

sys.setrecursionlimit(10000)

def leggi_configurazioni(percorso_file):
  input = open(percorso_file, "r")
  global variabile
  variabile = 5
  for line in input:
    if(line.startswith(";;;") or line == "\n"):
      continue
        
    parts = line.split(" = ")
    parts[1] = parts[1].replace("\n", "")
    globals()[parts[0]] = eval(parts[1])

def leggi_matrice(percorso_file):
    input = open(percorso_file, "r")
    A = []
    for line in input:
        if(line.startswith(";;;") or line == "\n"):
            continue
        ind = line.index("-")
        line = line[0:ind]
        parts = line.split(" ")[:-1]
        parts = [eval(i) for i in parts]
        A.append(parts)
    return A

def leggi_posizioni_arr_matrice(percorso_file):
    input = open(percorso_file, "r")
    A = list()
    N = 0
    M = 0
    for line in input:
        if (line.startswith(";;;") or line == "\n"):
            continue
        ind = line.index("-")
        line = line[0:ind]
        parts = line.split(" ")[:-1]
        M = len(parts)
        parts = [i for i, x in enumerate(parts) if x == "1"]

        A.append(parts)
        N = N + 1
    return A, N, M

def leggi_posizioni_arr_matrice_np(percorso_file):
    input = open(percorso_file, "r")
    A = []
    N = 0
    M = 0
    for line in input:
        if (line.startswith(";;;") or line == "\n"):
            continue
        ind = line.index("-")
        line = line[0:ind]
        parts = line.split(" ")[:-1]
        M = len(parts)
        parts = [i for i, x in enumerate(parts) if x == "1"]

        A.append(parts)
        N = N + 1
    A = np.array([np.array(xi, dtype=np.int16) for xi in A], dtype=object)
    return A, N, M

def actualsize(input_obj):
    memory_size = 0
    ids = set()
    objects = [input_obj]
    while objects:
        new = []
        for obj in objects:
            if id(obj) not in ids:
                ids.add(id(obj))
                memory_size += getsizeof(obj)
                new.append(obj)
        objects = gc.get_referents(*new)
    return memory_size

def convertSize(the_size):  
    if the_size < 1024:  
        return str(the_size) + " B"
    elif (the_size >= 1024) and (the_size < (1024 * 1024)):  
        return "%.2f KB" %(the_size/1024)  
    elif (the_size >= (1024 * 1024)) and (the_size < (1024 * 1024 * 1024)):  
        return "%.2f MB" %(the_size/(1024 * 1024))  
    else:  
        return "%.2f GB" %(the_size/(1024 * 1024 * 1024))
  
def stampa_output(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  try:
      with open(percorso, 'w') as f:
          f.write(";;;"+titolo+"\n")
          f.write(";;;matrice A\n")
          M = 0
          for i in range(len(A)):
              content = str(A[i])+" #indice:"+str(i)+"\n"
              content = content.replace("[", "")
              content = content.replace("]", "-")
              content = content.replace(",","")
              f.write(content)
              M = len(A[i])
          f.write("\n#Dimensioni: ")
          f.write("\nA = " +str(dim_A)+" bytes\n" )
          f.write("B = "+str(actualsize(B))+" bytes")
          f.write("\nCOV = " + str(actualsize(COV)) + " bytes\n")
          
          distribuzione_card_COV = [0 for element in range(M+1)]
          nuovo_set = set()
          f.write("\n#Risultato COV =\n")
          for i in range(len(COV)):
              content = str(COV[i])+"\n"
              content = content.replace("[(", "{")
              content = content.replace("(", "{")
              content = content.replace("]","}")
              content = content.replace("[","{")
              content = content.replace(")","}")
              content = content.replace("},",",")
              f.write(content)

              distribuzione_card_COV[len(COV[i])] += 1
              nuovo_set = nuovo_set.union(set(COV[i]))

          f.write("\nNumero delle partizioni trovate = " + str(len(COV)))
          
          f.write("\nDistribuzione della cardinalità del COV:")
          for i in range(0, len(distribuzione_card_COV)):
            if distribuzione_card_COV[i] != 0:
              f.write("\n\tNumero di elementi con cardinalità " + str(i) + " = " + str(distribuzione_card_COV[i]) )
          f.write("\nNumero di set = " + str(len(nuovo_set)))
          f.write("\nCardinalità = " + str(M))

          f.write("\n\n#Heap\n")
          
          size = 0
          f.write("Index Count     Size  Cumulative Size             Object Name\n")
          for row in heapStatus.stat.get_rows():
            
            if row.name != "str":
              if row.name == "list":
                row.count = row.count - 1
                row.size = row.size - getsizeof(riassunto_esecuzione)
              size += row.size
              f.write("%5d %5d %8d %8d %30s\n"%(row.index, row.count, row.size, size, row.name))
            

          f.write("\nSpazio totale di heap occupato = " +str(convertSize(size)))

          f.write("\n\nTempo di esecuzione: " + str(exec_time) + " millisecondi")   
          
          num_nodi_visitati = len(riassunto_esecuzione) - 1
          num_tot_nodi_visitati = sum(pow(2, i) for i in range(len(A)))
          perc_nodi_visitati = num_nodi_visitati / num_tot_nodi_visitati * 100
          f.write("\n\nNumero nodi visitati = " + str(num_nodi_visitati))
          f.write("\nNumero totale dei nodi visitati = " + str(num_tot_nodi_visitati))
          f.write("\nPercentuale dei nodi visitati sul totale = {0:5.4f}%\n".format(perc_nodi_visitati))

          leggi_configurazioni("config.txt")
          global elenco_nodi_visitati
          if(elenco_nodi_visitati):
            f.write("\n\nElenco dei nodi visitati:\n")
            for i in range(len(riassunto_esecuzione)):
              f.write(riassunto_esecuzione[i])

          print("Fine stampa")             
  except FileNotFoundError:
    print("errore")

def stampa_output_finale(percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  titolo = "RIASSUNTO ESECUZIONE COMPLETA"
  return stampa_output(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time)

def stampa_output_interrotto(percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  titolo = "RIASSUNTO ESECUZIONE INTERROTTA PER SUPERAMENTO LIMITE MASSIMO DI TEMPO"
  return stampa_output(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time)

def stampa_output_plus(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  try:
      with open(percorso, 'w') as f:
          f.write(";;;"+titolo+"\n")
          f.write(";;;matrice A\n")
          M = 0
          for i in range(len(A)):
              content = str(A[i])+" #indice:"+str(i)+"\n"
              content = content.replace("[", "")
              content = content.replace("]", "-")
              content = content.replace(",","")
              f.write(content)
              M = len(A[i])
          f.write("\n#Dimensioni: ")
          f.write("\nA = " +str(dim_A)+" bytes\n" )
          f.write("B = "+str(actualsize(B))+" bytes")
          f.write("\nCOV = "+ str(actualsize(COV)) + " bytes\n")

          elenco_visitati = list()
          elenco_cardinalita = list()

          switch = False
          for i in range(len(riassunto_esecuzione)):
            if riassunto_esecuzione[i] == "\nCardinalità di N:":
              switch = True

            if switch:
              elenco_cardinalita.append(riassunto_esecuzione[i])
            else:
              elenco_visitati.append(riassunto_esecuzione[i])

          if "\nCardinalità di N:" in riassunto_esecuzione:
            num_nodi_visitati = riassunto_esecuzione.index("\nCardinalità di N:")
          else:
            num_nodi_visitati = len(riassunto_esecuzione)
          num_tot_nodi_visitati = sum(pow(2, i) for i in range(len(A)))
          
          distribuzione_card_COV = [0 for element in range(M+1)]
          nuovo_set = set()
          f.write("\n#Risultato COV =\n")
          for i in range(len(COV)):
              content = str(COV[i])+"\n"
              content = content.replace("[(", "{")
              content = content.replace("(", "{")
              content = content.replace("]","}")
              content = content.replace("[","{")
              content = content.replace(")","}")
              content = content.replace("},",",")
              f.write(content)

              distribuzione_card_COV[len(COV[i])] += 1
              nuovo_set = nuovo_set.union(set(COV[i]))

          f.write("\nNumero delle partizioni trovate = " + str(len(COV)))
          
          f.write("\nDistribuzione della cardinalità del COV:")
          for i in range(0, len(distribuzione_card_COV)):
            if distribuzione_card_COV[i] != 0:
              f.write("\n\tNumero di elementi con cardinalità " + str(i) + " = " + str(distribuzione_card_COV[i]) )
          f.write("\nNumero di set = " + str(len(nuovo_set)))
          f.write("\nCardinalità = " + str(M)+"\n")

          for i in range(len(elenco_cardinalita)):
            f.write(elenco_cardinalita[i])

          f.write("\n\n#Heap\n")

          size = 0
          f.write("Index Count     Size  Cumulative Size             Object Name\n")
          for row in heapStatus.stat.get_rows():
            if row.name != "str":
              if row.name == "list":
                row.count = row.count - 1
                row.size = row.size - getsizeof(riassunto_esecuzione)
              size += row.size
              f.write("%5d %5d %8d %8d %30s\n"%(row.index, row.count, row.size, row.cumulsize, row.name))

          f.write("\nSpazio totale di heap occupato = " +str(convertSize(size)))

          f.write("\n\nTempo di esecuzione: " + str(exec_time) + " millisecondi")   

          perc_nodi_vistati = num_nodi_visitati / num_tot_nodi_visitati * 100
          f.write("\nNumero nodi visitati = " + str(num_nodi_visitati))
          f.write("\nNumero totale nodi visitati = " + str(num_tot_nodi_visitati))
          f.write("\nPercentuale nodi visitati = {0:5.4f}%\n".format(perc_nodi_vistati))

          leggi_configurazioni("config.txt")
          global elenco_nodi_visitati
          if(elenco_nodi_visitati):
            f.write("\n\nElenco dei nodi visitati:\n")
            for i in range(len(elenco_visitati)):
              f.write(elenco_visitati[i])

          print("Fine stampa")             
  except FileNotFoundError:
    print("errore")

def stampa_output_finale_plus(percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  titolo = "RIASSUNTO ESECUZIONE COMPLETA"
  return stampa_output_plus(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time)

def stampa_output_interrotto_plus(percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time):
  titolo = "RIASSUNTO ESECUZIONE INTERROTTA PER SUPERAMENTO LIMITE MASSIMO DI TEMPO"
  return stampa_output_plus(titolo, percorso, A, dim_A, B, riassunto_esecuzione, COV, heapStatus, exec_time)

def stampa_confronto_finale(percorso, A, dim_A, B_base, B_plus, riassunto_esecuzione_base, riassunto_esecuzione_plus, COV_base, COV_plus, heapStatus_base, heapStatus_plus, exec_time_base, exec_time_plus):
  titolo = "RIASSUNTO ESECUZIONE COMPLETA"
  return stampa_confronto(titolo, percorso, A, dim_A, B_base, B_plus, riassunto_esecuzione_base, riassunto_esecuzione_plus, COV_base, COV_plus, heapStatus_base, heapStatus_plus, exec_time_base, exec_time_plus)

def stampa_confronto_interrotto(percorso, A, dim_A, B_base, B_plus, riassunto_esecuzione_base, riassunto_esecuzione_plus, COV_base, COV_plus, heapStatus_base, heapStatus_plus, exec_time_base, exec_time_plus):
  titolo = "RIASSUNTO ESECUZIONE INTERROTTA PER SUPERAMENTO LIMITE MASSIMO DI TEMPO"
  return stampa_confronto(titolo, percorso, A, dim_A, B_base, B_plus, riassunto_esecuzione_base, riassunto_esecuzione_plus, COV_base, COV_plus, heapStatus_base, heapStatus_plus, exec_time_base, exec_time_plus)

def stampa_confronto(titolo, percorso, A, dim_A, B_base, B_plus, riassunto_esecuzione_base, riassunto_esecuzione_plus, COV_base, COV_plus, heapStatus_base, heapStatus_plus, exec_time_base, exec_time_plus):
  try:
      with open(percorso, 'w') as f:
          f.write(";;;"+titolo+"\n")
          f.write(";;;matrice A\n")
          for i in range(len(A)):
              content = str(A[i])+" #indice:"+str(i)+"\n"
              content = content.replace("[", "")
              content = content.replace("]", "-")
              content = content.replace(",","")
              f.write(content)
          f.write("\n#Dimensioni: ")
          f.write("\nA = " +str(dim_A)+" bytes\n\n" )
        
          f.write("B base = "+str(actualsize(B_base))+" bytes\n")
          f.write("B plus = "+str(actualsize(B_plus))+" bytes\n")

          f.write("\nCOV base = "+ str(actualsize(COV_base)) + " bytes\n")
          f.write("COV plus = "+ str(actualsize(COV_plus)) + " bytes\n")
          f.write("\n#Risultato COV base =\n")
          for i in range(len(COV_base)):
              content = str(COV_base[i])+"\n"
              content = content.replace("[(", "{")
              content = content.replace("(", "{")
              content = content.replace("]","}")
              content = content.replace("[","{")
              content = content.replace(")","}")
              content = content.replace("},",",")
              f.write(content)
          f.write("\n#Risultato COV plus =\n")
          for i in range(len(COV_plus)):
              content = str(COV_plus[i])+"\n"
              content = content.replace("[(", "{")
              content = content.replace("(", "{")
              content = content.replace("]","}")
              content = content.replace("[","{")
              content = content.replace(")","}")
              content = content.replace("},",",")
              f.write(content)

          f.write("\n#Heap base\n")

          size = 0
          f.write("Index Count     Size  Cumulative Size             Object Name\n")
          for row in heapStatus_base.stat.get_rows():
            if row.name != "str":
              if row.name == "list":
                row.count = row.count - 1
                row.size = row.size - getsizeof(riassunto_esecuzione_base)
              size += row.size
              f.write("%5d %5d %8d %8d %30s\n"%(row.index, row.count, row.size, size, row.name))

          f.write("\nSpazio totale di heap occupato = " +str(convertSize(size)))

          f.write("\n\n#Heap plus\n")

          size = 0
          f.write("Index Count     Size  Cumulative Size             Object Name\n")
          for row in heapStatus_plus.stat.get_rows():
            if row.name != "str":
              if row.name == "list":
                row.count = row.count - 1
                row.size = row.size - getsizeof(riassunto_esecuzione_plus)
              size += row.size
              f.write("%5d %5d %8d %8d %30s\n"%(row.index, row.count, row.size, size, row.name))

          f.write("\nSpazio totale di heap occupato = " +str(convertSize(size)))

          f.write("\n\nTempo di esecuzione base: " + str(exec_time_base) + " millisecondi")   
          f.write("\n\nTempo di esecuzione plus: " + str(exec_time_plus) + " millisecondi")  
          print("Fine stampa")             
  except FileNotFoundError:
    print("errore")

def ottieni_matrice_binaria(pos, N, M):
  A = np.zeros((N, M), dtype=int)
  for i in range(N):
    A[i][pos[i].tolist()] = 1
  return A

def stampa_storico(esito, versione, nome_file, dim_A, B, COV, riassunto_esecuzione, heapStatus, exec_time):
  percorso = './file/output/registro.log'
  try:
    with open(percorso, 'a') as f:

      size = 0

      for row in heapStatus.stat.get_rows():
        if row.name != "str":
          if row.name == "list":
            row.count = row.count - 1
            row.size = row.size - getsizeof(riassunto_esecuzione)
          size += row.size

      f.write("<" + esito + "> ")
      f.write("[" + versione + "]")
      f.write(" " + nome_file + "; ")
      f.write("Memoria = " + str(convertSize(size)) + "; ")
      f.write("Tempo = " + str(exec_time) + " ms; ")
      f.write("A = " + str(dim_A) + " B; ")
      f.write("B = " + str(actualsize(B)) + " B; ")
      f.write("COV = " + str(actualsize(COV)) + " B; ")
      f.write("Risultato = ")
      for i in range(len(COV)):
        f.write(str(COV[i]) + " - ")

      f.write("\n")
  except FileNotFoundError:
    print("errore")
  
  percorso = './file/output/registro.xlsx'
  workbook = xlsxwriter.Workbook(percorso)
  worksheet = workbook.add_worksheet()

  worksheet.set_column('A:H', 20)
  worksheet.set_column('D:E', 25)
  worksheet.set_column('I:I', 60)

  to_be = load_workbook(percorso)
  s = to_be.active
  last_row = len(list(s.rows))

  for row in range(1, last_row+1):
    for col in range(1, 10):
      worksheet.write(row-1, col-1, s.cell(row, col).value)

  row = last_row
  col = 0

  worksheet.write(row, col, esito)
  col = col + 1
  worksheet.write(row, col, versione)
  col = col + 1
  worksheet.write(row, col, nome_file)
  col = col + 1
  worksheet.write(row, col, str(convertSize(size)))
  col = col + 1
  worksheet.write(row, col, str(exec_time) + " ms")
  col = col + 1
  worksheet.write(row, col, str(dim_A) + " B")
  col = col + 1
  worksheet.write(row, col, str(actualsize(B)) + " B")
  col = col + 1
  worksheet.write(row, col, str(actualsize(COV)) + " B")
  col = col + 1

  stampa_esecuzione = ""
  for i in range(len(COV)):
    stampa_esecuzione = stampa_esecuzione + str(COV[i]) + " - "
  worksheet.write(row, col, stampa_esecuzione)
  
  workbook.close()

  wb = load_workbook(percorso)
  ws = wb.active
  ws.auto_filter.ref = ws.dimensions
  wb.save(percorso)